from flask import Flask, jsonify, render_template
from flask_restful import reqparse
import re
from openpyxl import load_workbook
import urllib

app = Flask(__name__)

parser = reqparse.RequestParser()
parser.add_argument('day', type=str, required=False)
parser.add_argument('program', type=str, required=False)
parser.add_argument('section', type=str, required=False)
parser.add_argument('course', type=str, required=False)
parser.add_argument('slot', type=int, required=False)
parser.add_argument('teacher', type=str, required=False)


def time_setter(t):
    if(t==1):
        return "8:00-9:00"
    elif(t==2):
        return "9:00-10:00"
    elif(t==3):
        return "10:00-11:00"
    elif(t==4):
        return "11:00-12:00"
    elif(t==5):
        return "12:00-1:00"
    elif(t==6):
        return "1:00-2:00"
    elif(t==7):
        return "2:00-3:00"
    elif(t==8):
        return "3:00-4:00"
    elif(t==9):
        return "4:00-5:00"
    

@app.route('/details/<string:course>')
def details(course):
    course = urllib.parse.unquote_plus(course)
    wb = load_workbook(filename='timetable.xlsx')
    l = []
    for sheet in wb.worksheets:
        if sheet.title=="Monday" or sheet.title=="Tuesday" or sheet.title=="Wednesday" or sheet.title=="Thursday" or sheet.title=="Friday" or sheet.title=="Saturday" or sheet.title=="Sunday":
            for i in range(1, sheet.max_row + 1):
                for j in range(sheet.max_column):
                    if (str(sheet[i][j].value).find("BCS") != -1 or str(sheet[i][j].value).find("BDS") != -1 or str(sheet[i][j].value).find("BAI") != -1 or str(sheet[i][j].value).find("BSE") != -1):
                        l.append({"Course Name": (" ".join(sheet[i][j].value.split()).upper().replace("/", "&")), "Day": (sheet.title.upper()), "Venue: ": (sheet[i][0].value.upper()), "Slot": (sheet[2][j].value), "Timings": time_setter((sheet[2][j].value))})
                        if str(sheet[i][j].value).find("Lab") != -1 or str(sheet[i][j].value).find("lab") != -1:
                            l.append({"Course Name": (" ".join(sheet[i][j].value.split()).upper().replace("/", "&")), "Day": (sheet.title.upper()), "Venue: ": (sheet[i][0].value.upper()), "Slot": (sheet[2][j].value)+1, "Timings": time_setter((sheet[2][j].value)+1)})
                            l.append({"Course Name": (" ".join(sheet[i][j].value.split()).upper().replace("/", "&")), "Day": (sheet.title.upper()), "Venue: ": (sheet[i][0].value.upper()), "Slot": (sheet[2][j].value)+2, "Timings": time_setter((sheet[2][j].value)+2)})

    
    r = []
    for x in l:
        if x["Course Name"] == course:
            r.append(x)
    return jsonify(r)
@app.route('/search')
def search():
    wb = load_workbook(filename='timetable.xlsx')
    program = str(parser.parse_args().get('program', None)).upper()
    section = str(parser.parse_args().get('section', None)).upper()
    day = str(parser.parse_args().get('day', None)).upper()
    teacher = str(parser.parse_args().get('teacher', None)).upper()
    slot = parser.parse_args().get('slot', None)
    course = str(parser.parse_args().get('course', None)).upper()
    
    s=[]
    if program!="NONE":
        s.append(program)
    if course!="NONE":
        s.append(course)
    if teacher!="NONE":
        s.append(teacher)
    if section!="NONE":
        s.append(section)

    l = []
    for sheet in wb.worksheets:
        if sheet.title=="Monday" or sheet.title=="Tuesday" or sheet.title=="Wednesday" or sheet.title=="Thursday" or sheet.title=="Friday" or sheet.title=="Saturday" or sheet.title=="Sunday":
            for i in range(1, sheet.max_row + 1):
                for j in range(sheet.max_column):
                    if (str(sheet[i][j].value).find("BCS") != -1 or str(sheet[i][j].value).find("BDS") != -1 or str(sheet[i][j].value).find("BAI") != -1 or str(sheet[i][j].value).find("BSE") != -1):
                        l.append({"Course Name": (" ".join(sheet[i][j].value.split()).upper().replace("/", "&")), "Day": (sheet.title.upper()), "Venue: ": (sheet[i][0].value.upper()), "Slot": (sheet[2][j].value), "Timings": time_setter((sheet[2][j].value))})
                        if str(sheet[i][j].value).find("Lab") != -1 or str(sheet[i][j].value).find("lab") != -1:
                            l.append({"Course Name": (" ".join(sheet[i][j].value.split()).upper().replace("/", "&")), "Day": (sheet.title.upper()), "Venue: ": (sheet[i][0].value.upper()), "Slot": (sheet[2][j].value)+1, "Timings": time_setter((sheet[2][j].value)+1)})
                            l.append({"Course Name": (" ".join(sheet[i][j].value.split()).upper().replace("/", "&")), "Day": (sheet.title.upper()), "Venue: ": (sheet[i][0].value.upper()), "Slot": (sheet[2][j].value)+2, "Timings": time_setter((sheet[2][j].value)+2)})
    r = []
    for x in l:
        print(s)
        print(day)
        print(slot)
        if day!="NONE" and x["Day"]!=day:
            continue
        if slot is not None and x["Slot"]!=slot:
            continue
        if all(y in x["Course Name"] for y in s):
            r.append(x)
        
    return jsonify(r)

@app.route('/names')
def names():
    wb = load_workbook(filename='timetable.xlsx')
    l = []
    for sheet in wb.worksheets:
        if sheet.title=="Monday" or sheet.title=="Tuesday" or sheet.title=="Wednesday" or sheet.title=="Thursday" or sheet.title=="Friday" or sheet.title=="Saturday" or sheet.title=="Sunday":
            for i in range(1, sheet.max_row + 1):
                for j in range(sheet.max_column):
                    if str(sheet[i][j].value).find("BCS") != -1 or str(sheet[i][j].value).find("BDS") != -1 or str(sheet[i][j].value).find("BAI") != -1 or str(sheet[i][j].value).find("BSE") != -1:
                        l.append({"Course Name": (" ".join(sheet[i][j].value.split()).upper().replace("/", "&")), "Day": (sheet.title.upper()), "Venue: ": (sheet[i][0].value.upper()), "Slot": (sheet[2][j].value)})
    r = []
    for x in l:
        if x["Course Name"] not in r:
            r.append(x["Course Name"])
    return jsonify(r)

@app.route('/navigate')
def nav():
    wb = load_workbook(filename='timetable.xlsx')
    l = []
    dupli=[]
    for sheet in wb.worksheets:
        if sheet.title=="Monday" or sheet.title=="Tuesday" or sheet.title=="Wednesday" or sheet.title=="Thursday" or sheet.title=="Friday" or sheet.title=="Saturday" or sheet.title=="Sunday":
            for i in range(1, sheet.max_row + 1):
                for j in range(sheet.max_column):
                    if (str(sheet[i][j].value).find("BCS") != -1 or str(sheet[i][j].value).find("BDS") != -1 or str(sheet[i][j].value).find("BAI") != -1 or str(sheet[i][j].value).find("BSE") != -1) and (" ".join(sheet[i][j].value.split()).upper().replace("/", "&")) not in dupli:
                        l.append({"Course Name": (" ".join(sheet[i][j].value.split()).upper().replace("/", "&")), "Day": (sheet.title.upper()), "Venue: ": (sheet[i][0].value.upper()), "Slot": (sheet[2][j].value)})

    course_names = [[]]

    for x in l:
        if x["Course Name"] not in course_names:
            course_names.append([x["Course Name"], "details/" + urllib.parse.quote_plus(x["Course Name"])])
    
    return render_template("index.html", course_names=course_names)


