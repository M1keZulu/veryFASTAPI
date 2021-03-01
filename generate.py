import json
from openpyxl import load_workbook

def time_setter(t):
    if(t==1):
         return "8:00-9:00"
    elif(t==2):
        return "9:00-10:00"
    elif(t==3):
        return "10:00-11:00"
    elif(t==4):
        return "11:00-12:00"
    elif(t==5):
        return "12:00-1:00"
    elif(t==6):
        return "1:00-2:00"
    elif(t==7):
        return "2:00-3:00"
    elif(t==8):
        return "3:00-4:00"
    elif(t==9):
        return "4:00-5:00"

def gen():
    wb = load_workbook(filename='timetable.xlsx')
    l = []
    for sheet in wb.worksheets:
        if sheet.title=="Monday" or sheet.title=="Tuesday" or sheet.title=="Wednesday" or sheet.title=="Thursday" or sheet.title=="Friday" or sheet.title=="Saturday" or sheet.title=="Sunday":
            for i in range(1, sheet.max_row + 1):
                for j in range(sheet.max_column):
                    if (str(sheet[i][j].value).find("BCS") != -1 or str(sheet[i][j].value).find("BDS") != -1 or str(sheet[i][j].value).find("BAI") != -1 or str(sheet[i][j].value).find("BSE") != -1):
                        l.append({"Course Name": (" ".join(sheet[i][j].value.split()).upper().replace("/", "&")), "Day": (sheet.title.upper()), "Venue": (sheet[i][0].value.upper()), "Slot": (sheet[2][j].value), "Timings": time_setter((sheet[2][j].value))})
                        if str(sheet[i][j].value).find("Lab") != -1 or str(sheet[i][j].value).find("lab") != -1:
                            l.append({"Course Name": (" ".join(sheet[i][j].value.split()).upper().replace("/", "&")), "Day": (sheet.title.upper()), "Venue": (sheet[i][0].value.upper()), "Slot": (sheet[2][j].value)+1, "Timings": time_setter((sheet[2][j].value)+1)})
                            l.append({"Course Name": (" ".join(sheet[i][j].value.split()).upper().replace("/", "&")), "Day": (sheet.title.upper()), "Venue": (sheet[i][0].value.upper()), "Slot": (sheet[2][j].value)+2, "Timings": time_setter((sheet[2][j].value)+2)})
    with open("data.txt", "w") as f:
        json.dump(l, f)

gen()